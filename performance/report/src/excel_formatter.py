"""
Excel 格式化工具

功能：
1. 将 DataFrame 中格式化后的字符串（如 "12.34%"、"1,234"）解析为数值，
   并写入对应的 Excel 数字格式，使 WPS / Excel 可以正确识别为数字。
2. 自动计算列宽（按内容最长文本适配，含标题行）。
3. 为标题行添加加粗、居中、浅灰背景。
4. 统一设置行高，避免 WPS 中行高压缩。
"""

from __future__ import annotations

import re
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── 正则：匹配已格式化的数值字符串 ──────────────────────────────────

# 百分数：可选负号，数字部分，小数点后若干位，结尾 %
_RE_PCT = re.compile(r"^(-?\d+(?:\.\d+)?)%$")

# 千分位整数/小数（如 "1,234" 或 "1,234.56"），可带负号
_RE_COMMA = re.compile(r"^-?[\d,]+(?:\.\d+)?$")

# 普通小数 / 整数（无逗号、无 %）
_RE_NUM = re.compile(r"^-?\d+(?:\.\d+)?$")


def _decimal_places(s: str) -> int:
    """从字符串中提取小数点后的位数（保留尾部零）。"""
    if "." in s:
        return len(s.split(".")[1])
    return 0


def _parse_cell(value: Any) -> tuple[Any, str | None]:
    """
    尝试将字符串单元格值解析为数值，并返回对应 Excel 数字格式字符串。

    Returns:
        (parsed_value, excel_format) — 若无法解析，excel_format 为 None。
    """
    if not isinstance(value, str):
        return value, None

    s = value.strip()
    if not s or s in ("N/A", "-"):
        return value, None

    # ── 百分数 ──────────────────────────────────────────────
    m = _RE_PCT.match(s)
    if m:
        try:
            num = float(m.group(1)) / 100.0
            # 根据小数位数决定格式精度
            dec = _decimal_places(m.group(1))
            if dec >= 4:
                fmt = "0.0000%"
            elif dec >= 2:
                fmt = "0.00%"
            elif dec == 1:
                fmt = "0.0%"
            else:
                fmt = "0%"
            return num, fmt
        except ValueError:
            return value, None

    # ── 千分位数字（含小数） ─────────────────────────────────
    if "," in s:
        m2 = _RE_COMMA.match(s)
        if m2:
            try:
                clean = s.replace(",", "")
                if "." in clean:
                    num = float(clean)
                    dec = len(clean.split(".")[1])
                    fmt = f'#,##0.{"0" * dec}' if dec > 0 else "#,##0"
                else:
                    num = int(clean)
                    fmt = "#,##0"
                return num, fmt
            except ValueError:
                return value, None

    # ── 普通数字 ─────────────────────────────────────────────
    m3 = _RE_NUM.match(s)
    if m3:
        try:
            if "." in s:
                num = float(s)
                dec = _decimal_places(s)
                if dec >= 4:
                    fmt = "0.0000"
                elif dec >= 2:
                    fmt = "0.00"
                elif dec == 1:
                    fmt = "0.0"
                else:
                    fmt = "0"
            else:
                num = int(s)
                fmt = "0"
            return num, fmt
        except ValueError:
            return value, None

    return value, None


# ── 估算单元格显示宽度（字符数）──────────────────────────────────────

def _display_width(val: Any) -> float:
    """
    估算内容显示宽度（以 Excel 字符单位计）。
    中文字符约占 2 个英文字符宽度。
    """
    if val is None:
        return 1.0
    s = str(val)
    width = 0.0
    for ch in s:
        if "\u4e00" <= ch <= "\u9fff" or "\u3000" <= ch <= "\u303f":
            width += 2.2   # 中文字符更宽
        else:
            width += 1.1
    return width


# ── 主格式化函数 ──────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")   # 浅灰标题背景
_HEADER_FONT = Font(bold=True, name="微软雅黑", size=10)
_DATA_FONT   = Font(name="微软雅黑", size=10)
_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=False)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 列宽上下限（Excel 字符单位）
_COL_MIN = 6
_COL_MAX = 35


def write_excel(
    sheets: dict[str, pd.DataFrame],
    output_path,
) -> None:
    """
    将多个 DataFrame 写入 Excel，每张表一个 Sheet，并应用格式化。

    Args:
        sheets:      {sheet_name: df} 有序字典
        output_path: 输出路径（pathlib.Path 或字符串）

    实现说明：
      - 单遍扫描完成写入 + 样式设置（合并 raw / format_sheet 两次扫描）
      - ``ws.append`` 一行写入比逐 cell ``ws.cell(...)`` 快约 3-5×
      - 单元格样式用 module-level 单例（Font / Alignment / PatternFill）共享
      - ``_parse_cell`` / ``_display_width`` 对**字符串**做 LRU 缓存，因为同一
        sheet 内大量字符串重复（例：周度_品种 24011 行 × 16 列里有大量重复的
        "0.00%"、"N/A"、品种代码等）
    """
    import openpyxl
    from functools import lru_cache

    # 全局 LRU 缓存：解析 + 宽度计算（命中率极高）
    @lru_cache(maxsize=65536)
    def _parse_str(s: str) -> tuple:
        v, fmt = _parse_cell(s)
        return v, fmt

    @lru_cache(maxsize=65536)
    def _width_str(s: str) -> float:
        return _display_width(s)

    def _parse_any(val):
        if isinstance(val, str):
            return _parse_str(val)
        return val, None

    def _width_any(val) -> float:
        if val is None:
            return 1.0
        if isinstance(val, str):
            return _width_str(val)
        return _display_width(val)

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        n_rows, n_cols = df.shape

        col_names = list(df.columns)
        # 先写表头（一次 append）
        ws.append(col_names)
        # 表头格式
        for col_idx in range(1, n_cols + 1):
            c = ws.cell(row=1, column=col_idx)
            c.font      = _HEADER_FONT
            c.fill      = _HEADER_FILL
            c.alignment = _HEADER_ALIGN
        ws.row_dimensions[1].height = 36

        # 预解析所有数据（vectorize-able 的部分用 numpy 提速）
        # df.iat[i,j] 比 df.values[i,j] 慢得多；先 to_numpy 一次
        arr = df.to_numpy(dtype=object, copy=False)

        col_widths: list[float] = [max(_COL_MIN, _width_any(c)) for c in col_names]

        # 数据：逐行 append；append 内部把整行批量写入比逐 cell 快
        for row_idx in range(n_rows):
            row_vals = []
            row_fmts: list[tuple[int, str]] = []
            for col_idx in range(n_cols):
                raw = arr[row_idx, col_idx]
                parsed_val, fmt = _parse_any(raw)
                row_vals.append(parsed_val)
                if fmt is not None:
                    row_fmts.append((col_idx + 1, fmt))
                # 列宽
                w = _width_any(raw)
                if w > col_widths[col_idx]:
                    col_widths[col_idx] = w
            ws.append(row_vals)
            excel_row = row_idx + 2
            # 行级样式：只对该行已写入的 cell 设置
            for col_idx in range(1, n_cols + 1):
                c = ws.cell(row=excel_row, column=col_idx)
                c.font      = _DATA_FONT
                c.alignment = _CENTER
            for col_idx, fmt in row_fmts:
                ws.cell(row=excel_row, column=col_idx).number_format = fmt
            ws.row_dimensions[excel_row].height = 18

        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max(width + 1.5, _COL_MIN), _COL_MAX,
            )
        ws.freeze_panes = "A2"

    wb.save(output_path)
